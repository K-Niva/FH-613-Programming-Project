import os
import csv
from flask import Flask, render_template, request, send_from_directory, Response
import openpyxl
import requests
from werkzeug.utils import secure_filename
import json

UPLOAD_FOLDER = 'uploads'
DOWNLOAD_FOLDER = 'downloads'
ALLOWED_EXTENSIONS = {'xlsx', 'csv'}

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER
app.secret_key = 'supersecretkey'

def allowed_file(filename):
    """Checks if the file extension is allowed."""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
if not os.path.exists(DOWNLOAD_FOLDER):
    os.makedirs(DOWNLOAD_FOLDER)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return Response("No file part", status=400)

    file = request.files['file']

    if file.filename == '':
        return Response("No selected file", status=400)

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        if filename.endswith('.xlsx'):
            return Response(process_excel_stream(filepath, filename), mimetype='text/event-stream')
        elif filename.endswith('.csv'):
            return Response(process_csv_stream(filepath, filename), mimetype='text/event-stream')
        else:
            return Response("Invalid file type", status=400)
    else:
        return Response("Allowed file types are .xlsx and .csv", status=400)


def process_csv_stream(filepath, original_filename):
    """Processes a .csv file and yields progress updates."""
    
    def generate():
        processed_rows = []
        
        # First, count the number of URLs to process
        with open(filepath, 'r', newline='', encoding='utf-8') as infile:
            reader = csv.reader(infile)
            try:
                header = next(reader)
                # Count rows that have a value in the first column
                total_urls = sum(1 for row in reader if row and row[0])
            except StopIteration:
                total_urls = 0

        yield f'data: {json.dumps({"message": f"Found {total_urls} URLs in {original_filename}"})}\n\n'
        yield f'data: {json.dumps({"total": total_urls})}\n\n'

        checked_count = 0
        with open(filepath, 'r', newline='', encoding='utf-8') as infile:
            reader = csv.reader(infile)
            try:
                header = next(reader)
                if "HTTP Status" not in header:
                    header.append("HTTP Status")
                processed_rows.append(header)

                for row in reader:
                    if row and row[0]:
                        checked_count += 1
                        url = row[0].strip()
                        if url.startswith(('http://', 'https://')):
                            try:
                                response = requests.get(url, timeout=10, allow_redirects=True)
                                status_code = response.status_code
                            except requests.exceptions.RequestException:
                                status_code = "Error - Unreachable"
                            row.append(status_code)
                        else:
                            row.append("Not a valid URL")
                        yield f'data: {json.dumps({"checked": checked_count})}\n\n'
                    else:
                        row.append("") # Append empty status for empty URL rows
                    processed_rows.append(row)

            except StopIteration:
                pass

        base, ext = os.path.splitext(original_filename)
        output_filename = f"{base}_processed{ext}"
        output_filepath = os.path.join(app.config['DOWNLOAD_FOLDER'], output_filename)
        with open(output_filepath, 'w', newline='', encoding='utf-8') as outfile:
            writer = csv.writer(outfile)
            writer.writerows(processed_rows)

        yield f'data: {json.dumps({"done": True, "filename": output_filename})}\n\n'

    return generate()

def process_excel_stream(filepath, original_filename):
    """Processes an .xlsx file and yields progress updates."""
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    
    def generate():
        # Total rows minus the header row
        total_urls = sheet.max_row - 1 if sheet.max_row > 1 else 0

        yield f'data: {json.dumps({"message": f"Found {total_urls} URLs in {original_filename}"})}\n\n'
        yield f'data: {json.dumps({"total": total_urls})}\n\n'

        # Find or create the "HTTP Status" column
        status_column = sheet.max_column + 1
        header_found = False
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == "HTTP Status":
                status_column = col
                header_found = True
                break
        if not header_found:
            sheet.cell(row=1, column=status_column, value="HTTP Status")

        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=1)
            if cell.value:
                url = str(cell.value).strip()
                if url.startswith(('http://', 'https://')):
                    try:
                        response = requests.get(url, timeout=10, allow_redirects=True)
                        status_code = response.status_code
                    except requests.exceptions.RequestException:
                        status_code = "Error - Unreachable"
                    sheet.cell(row=cell.row, column=status_column, value=status_code)
                else:
                    sheet.cell(row=cell.row, column=status_column, value="Not a valid URL")
            
            # Send progress update for each row processed
            yield f'data: {json.dumps({"checked": row_index - 1})}\n\n'

        base, ext = os.path.splitext(original_filename)
        output_filename = f"{base}_processed{ext}"
        output_filepath = os.path.join(app.config['DOWNLOAD_FOLDER'], output_filename)
        workbook.save(output_filepath)
        yield f'data: {json.dumps({"done": True, "filename": output_filename})}\n\n'
        
    return generate()


@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(app.config['DOWNLOAD_FOLDER'], filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)